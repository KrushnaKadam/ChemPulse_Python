#!/usr/bin/env python
# coding: utf-8

# In[1]:
 
import uvicorn,fastapi
from fastapi import FastAPI
import pandas as pd
import numpy as np
import math,json,os,pycountry,pathlib
import datetime
from azure.storage.blob import BlobServiceClient, ContentSettings,BlobClient
import os, uuid, sys
from azure.identity import DefaultAzureCredential
from azure.storage.filedatalake import DataLakeServiceClient
from azure.core._match_conditions import MatchConditions
from azure.storage.filedatalake._models import ContentSettings
import pandas as pd
import pyodbc
import pandas as pd
import pyodbc
import urllib.parse
import urllib
import sqlalchemy
import datetime
import pandas as pd
import requests
import json,re,os
from pathlib import Path
from sqlalchemy import create_engine
import numpy as np
from scipy import stats
import os.path
import datetime
import ast
import openpyxl
from openpyxl.styles import PatternFill

start_timestamp=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3].replace(':',"_").replace('.',"_").replace(' ',"_")
_connectionString='''DefaultEndpointsProtocol=https;AccountName=gtoadlsdev;AccountKey=Dm4zudPhZ7xJS9pCHkExFykR1ZZq8U1Rcc6YkuYiEUGmpjAcKxD3T0Ze/M6of4ATBmLXDNuGCv9DxBaI7GxHIQ==;EndpointSuffix=core.windows.net'''
_containerName="chempulse-container"
blob_service_client = BlobServiceClient.from_connection_string(_connectionString)
# Instantiate a new ContainerClient
container_client = blob_service_client.get_container_client(_containerName)
currency=['SVC','BGN','MKD','UYI','VND','GNF','NOK','DOP','JOD','SZL','MAD','DZD','XAF','GMD','AWG','KYD','CZK','SSP','KHR','MXN','TZS','CHF','IDR','NIO','COP','USN','MOP','ANG','NPR','SYP','BDT','LAK','ETB','BTN','CVE','PHP','MGA','PGK','LSL','TOP','ARS','SCR','BYN','GIP','KWD','SLE','RSD','USD','XDR','IRR','ILS','BND','FJD','RWF','ISK','COU','DJF','BZD','VUV','ERN','RON','ZAR','TWD','BRL','TND','MRU','GBP','VEF','MUR','MDL','DKK','VED','LYD','JPY','RUB','HNL','GYD','TTD','BOV','CRC','MWK','QAR','KPW','HRK','SOS','EGP','KZT','NZD','XPF','UGX','EUR','BAM','AUD','AZN','CHE','PLN','BWP','WST','XSU','MZN','UAH','BOB','AOA','BHD','HKD','NGN','BIF','UYU','UZS','JMD','MNT','OMR','SDG','GHS','XCD','MXV','AMD','CNY','BMD','LRD','FKP','SRD','SAR','SGD','GEL','CUC','XOF','CLF','PEN','INR','BBD','STN','YER','CUP','PYG','CHW','SHP','KMF','MMK','PKR','IQD','XUA','HTG','KGS','GTQ','ZMW','AED','KRW','SBD','PAB','SEK','TMT','TJS','BSD','HUF','CAD','NAD','TRY','ZWL','MYR','KES','LKR','ALL','MVR','CDF','LBP','CLP','THB']

mapping = { country.alpha_2:  country.name.upper() for country in pycountry.countries}
app = FastAPI()

# In[2]:


user_name='gtoadmin'
password='Admin@gt0' 
Database_Name='gtosqldbdev' 
Server_Name='gtosqlserverdev.database.windows.net'

class database:
    def __init__(self , SQLUser , SQLPassword , DatabaseName , ServerName ):
        driver="{ODBC Driver 17 for SQL Server}"
        self.conn = pyodbc.connect(f'''DRIVER={driver}; Server={ServerName }; 
                                UID={SQLUser}; PWD={SQLPassword}; DataBase={DatabaseName}''')
    
        server = ServerName
        database = DatabaseName
        user = SQLUser
        password = f"{SQLPassword}"
        driver = "{ODBC Driver 17 for SQL Server}"
        params = urllib.parse.quote_plus(f"""Driver={driver};
                                        Server=tcp:{server},1433;
                                        Database={database};
                                        Uid={user};Pwd={password};
                                        Encrypt=yes;
                                        TrustServerCertificate=no;
                                        Connection Timeout=30;""")
        self.conn_str = 'mssql+pyodbc:///?autocommit=true&odbc_connect={}'.format(
            params)
        
    # Fetching the data from the selected table using SQL query
    def read_table(self,query):
        RawData= pd.read_sql_query(query, self.conn)
        return RawData
    
    def execute_query(self,query):
        cursor=self.conn.cursor()
        cursor.execute(query)
        cursor.commit()
        
    def close_conn_eng(self):
        self.conn.close()
    def insert_data(self, df, table_name ,schema_name):
        try :
            self.engine1 = create_engine(self.conn_str, fast_executemany=True)
            df.to_sql(table_name, con = self.engine1 ,if_exists='append', index=False,schema=schema_name)
        except Exception as e :
            print(e)    
            
new_dbobj=database(user_name , password , Database_Name , Server_Name)


# In[3]:


class Adls:
    '''adls class to
    connect to adls gen2
    get directory object
    create sub directories
    upload files in subdirectories'''

    def __init__(self, storage_account_name: str, storage_account_key: str):
        '''
        azure data lake storage gen2 class 
        storage_account_name : name of account 
        storage_account_key : access key for storage account '''
        # self.logger = logger
        self.storage_account_name = storage_account_name
        self.storage_account_key = storage_account_key

    def initialize_storage_account(self):
        '''
        initialise datalake sevice client 
        params : storage_account_name , storage_account_key'''
        try:
            # global service_client
            service_client = DataLakeServiceClient(
                account_url="{}://{}.dfs.core.windows.net".format(
                    "https", self.storage_account_name),
                credential=self.storage_account_key)
            return service_client
        except Exception as e:
            print('error in initializing data lake service client ',e)

    def get_container_directory(self, container: str, directory: str, service_client):
        '''returns container directory object 
        params : container object , directory object , datalake service client'''
        # file_system="gtoprimedev"
        try:
            file_system_client = service_client.get_file_system_client(
                container)

            directory_client = file_system_client.get_directory_client(
                directory)

            return directory_client
        except Exception as e:
            print('error in returning directory object',e)

    def create_sub_dir(self,directory_client, sub_dirname):
        '''create subdirectory inside given directory '''
        try:
            subdir_client = directory_client.create_sub_directory(sub_dirname)

            return subdir_client
        except Exception as e:
            print('error in creating sub-directory object',e)

    def rename_directory(self, rename_dir, directory_client):
        '''renames a directory '''
        try:
            new_dir_name = rename_dir
            directory_client.rename_directory(
                new_name=directory_client.file_system_name + '/' + new_dir_name)
        except Exception as e:
            print(e)

    def upload_file_to_subdirectory(self, text, subdir_client, file_name):
        '''
        uploads file to subdirectory '''
        try:
            file_client = subdir_client.create_file(file_name)
            file_client.append_data(data=text, offset=0, length=len(text))
            file_client.flush_data(len(text))
            print(f'uploaded file  : {file_name}')
            # return
        except Exception as e:
            print(e)
            pass

    def dir_exists(self, directory_client):
        ''' checks if a directory exists and 
            returns a boolean '''
        result = directory_client.exists()
        return result

    def delete_directory(self, directory_client):
        ''' deletes a directory '''
        try:
            directory_client.delete_directory()
            print('directory deleted ')
        except Exception as e:
            print(f'error while deleting directory',e)

    def download_file_from_directory(self , directory_client, file_path, file_name ):
        try:
            # file_system_client = service_client.get_file_system_client(file_system="my-file-system")
            # directory_client = file_system_client.get_directory_client("my-directory")
            local_file = open(file_path,'wb')
            file_client = directory_client.get_file_client(file_name)
            download = file_client.download_file()
            downloaded_bytes = download.readall()
            local_file.write(downloaded_bytes)
            local_file.close()
        except Exception as e:
            print(e)
        
    def list_directory_contents(self, file_system_client, directory):
        try:
            paths = file_system_client.get_paths(path= directory)
            for path in paths:
                print(path.name + '\n')
        except Exception as e:
            print(e)

    def upload_file_to_directory( self, directory_client, local_filepath, file_name ):
        try:
            file_client = directory_client.create_file(file_name)
            local_file = open(local_filepath,'r')
            print('Read_done')
            file_contents = local_file.read()
            file_client.append_data(data=file_contents, offset=0, length=len(file_contents))

            file_client.flush_data(len(file_contents))

        except Exception as e:
            print(e)

def read_file_adls(filepath,local_file_path_to_save,Adls_target_folder):
    # credentials
    ADLSName = "gtoadlsdev"
    StorageAccountKey = "Dm4zudPhZ7xJS9pCHkExFykR1ZZq8U1Rcc6YkuYiEUGmpjAcKxD3T0Ze/M6of4ATBmLXDNuGCv9DxBaI7GxHIQ=="

    #variables
    ContainerName = "chempulse-container"
    downloadFolder = f'{Adls_target_folder}/Input'#"RFP/Input"
    
    localfilepath = local_file_path_to_save#os.path.join(temp_local_path, "localfile.csv") #f"{os.getcwd()}/localfile.csv"
    fileName = filepath#"2023_03_10_GE_LYB.xlsx" 

    # create obj for ADLS class
    AdlsObject = Adls( ADLSName , StorageAccountKey )
    ServiceClient = AdlsObject.initialize_storage_account()
    # get container client 
    file_system_client1 = ServiceClient.get_file_system_client(ContainerName)
    # getting download directory client
    download_directory_client = AdlsObject.get_container_directory( ContainerName , downloadFolder , ServiceClient)

    # download file from directory
    AdlsObject.download_file_from_directory( download_directory_client, localfilepath,fileName )
    
def upload_excel(_file,_connectionString,_containerName,_fileName):            
    try:  
        blob = BlobClient.from_connection_string(conn_str=_connectionString,
                                                container_name=_containerName,
                                                blob_name=_fileName) 
                    
        # Instantiate a new BlobClient            
        blob_client = container_client.get_blob_client(blob=_fileName)
        # upload data
        # Upload the created file
        with open(_file, "rb") as data:
            blob_client.upload_blob(data)
    except Exception as e:
            print(e)


# In[4]:


#2
def is_valid_expected_lead_time_warning(num):
    try:
        try:
            if math.isnan(num):
                return True
        except:
            flag=0
        if str(num).split('.')[0].isnumeric() or num is None or num is np.nan or 'month' in str(num).lower() or 'day' in str(num).lower():
            return True
        return False
    except:
        return False

#3
def is_valid_can_you_fulfil_100_of_demand(num):
    try:
        try:
            if math.isnan(num):
                return True
        except:
            flag=0
        if str(num).replace('%','').split('.')[0].isnumeric() or num is None or num is np.nan or 'y'== str(num).lower().strip() or 'n'== str(num).lower().strip() or '%' in str(num).lower().strip() or str(num).strip().lower() in ['yes','no','0','1']:
            return True
        return False
    except:
        return False

#4,5
def is_valid_numeric_warning(num):
    try:
        try:
            if math.isnan(num):
                return True
        except:
            flag=0
        if str(num).isnumeric() or num is None or num is np.nan or str(num).replace('%','').split('.')[0].isnumeric() :
            return True
        return False
    except:
        return False
    
#8,9
def is_valid_hcb_limit_warning(num):
    try:
        try:
            if math.isnan(num):
                return True
        except:
            flag=0
        if num is None or num is np.nan or str(num).strip().lower() in ['yes','no','0','1'] or str(num).split('.')[0] in ['0','1']:
            return True
        return False
    except:
        return False
    
#10
#2.2,2.3
def is_valid_alternate_product_pricing(num):
    try:
        try:
            if math.isnan(num):
                return True
        except:
            flag=0
        
        try:
            if type(int(num))==int:
                return True
        except:
            flag=0
        
        
        if str(num).replace('-','').split('.')[0].isnumeric() or num is None or num is np.nan:
            return True
        return False
    except:
        return False


def is_valid_numeric(num):
    try:
        if math.isnan(num) or str(int(num)).isnumeric() or num is None or num is np.nan :
            return True
        return False
    except:
        return False
    
def is_currency(curr):
    for currency in curr:
        if currency.lower() in curr.lower():
            return True
    return False
    
def is_valid_currency(currency_str):
    try:
        try:
            if math.isnan(currency_str):
                return True
        except:
            flag=0
        
        if is_valid_currency(currency_str):
            return True
        
        if currency_str.upper().strip() in currency or currency_str is None or currency_str is None or currency_str is np.nan or currency_str.strip()=='N/A' or pd.isnull(currency_str):
            return True
        else:
            return False
    except:
        return False


def is_valid_bid_price_base(base):
    try:
        try:
            if math.isnan(base):
                return True
        except:
            flag=0
        if base.upper().strip() in ['FORMULA','FIXED - QUARTERLY'] or base is None or base is np.nan or pd.isnull(base):
            return True
        else:
            return False
    except:
        return False
    

def is_valid_country(country):
    try:
        try:
            if math.isnan(country):
                return True
        except:
            flag=0
        if str(country).strip().upper() in mapping or country is None or country is np.nan or pd.isnull(country) :
            return True
        elif str(country).strip().upper() in mapping.values():
            return True
        return False
    except:
        return False
    

def update_Can_you_fulfil_100_demand(val):
    try:
        if str(val).lower().strip()=='n' or str(val).lower().strip()=='no' or str(val).lower().strip()=='0':
            return 0
        elif str(val).lower().strip()=='y' or str(val).lower().strip()=='yes' or str(val).lower().strip()=='1':
            return 1
        return None
    except:
        return None
    
def standardize_Expected_Lead_Time(val):
    try:
        month_flag=0
        monthh={"one":1, "two":2, "three":3, "four":4, "five":5, "six":6, "seven":7, "eight":8,
        "nine":9, "ten":10, "eleven":11, "twelve":12}
        digit_words=list(monthh.keys())
        
        val=str(val).lower().strip()
        if 'month' in val:
            month_flag=1
        val=val.replace('month','').replace('day','').replace('s','')
        for word in digit_words:
            if word in val:
                val=val.replace(word,str(monthh[word]))
        if month_flag==1:
            return int(val)*30
        else:
            return int(val)
    except:
        return None    

def get_adls_folder(file_type_id):
    file_type_id=str(file_type_id).strip()
    if file_type_id=='1':
        return 'CommodityPricing'
    elif file_type_id=='2':
        return 'BaseData'
    elif file_type_id=='3':
        return 'RFP'
        
def create_local_folder():
    path = os.path.join(pathlib.Path().resolve(), "file")
    # Check whether the specified path exists or not
    isExist = os.path.exists(path)
    if not isExist:
        # Create a new directory because it does not exist
        os.makedirs(path)
        print("The new directory is created!")
        
#2.4
def is_not_null(val):
    try:
        try:
            if math.isnan(val):
                return False
        except:
            flag=0
        if val is None or val is np.nan or str(val).strip() =='':
            return False
        return True
    except:
        return False
        
#2.5
def is_only_char_string(val):
    try:
        try:
            if math.isnan(val):
                return True
        except:
            flag=0
        if val is None or val is np.nan or str(val).strip() =='':
            return True
        
        if any(chr.isdigit() for chr in val):
            return False
        return True
    except Exception as e:
        print(e,val)
        



# In[5]:


MAX = 50

# Function to print Excel column name
# for a given column number
def printString(n):

    # To store result (Excel column name)
    n=int(n)
    string = ["\0"] * MAX

    # To store current index in str which is result
    i = 0

    while n > 0:
        # Find remainder
        rem = n % 26

        # if remainder is 0, then a
        # 'Z' must be there in output
        if rem == 0:
            string[i] = 'Z'
            i += 1
            n = (n / 26) - 1
        else:
            string[i] = chr((rem - 1) + ord('A'))
            i += 1
            n = n // 26
    string[i] = '\0'

    # Reverse the string and print result
    string = string[::-1]
    return("".join(string))


# In[8]:


def remove_empty_space(val):
    new_val=""
    cnt=0
    for char in val:
        if char.isdigit() or char.isalpha():
            new_val+=char
            cnt+=1
    return new_val
            

def processing(file_path):
    xls = pd.ExcelFile(file_path)
    raw_rfp_file = pd.read_excel(xls, 'Proposal Sheet')
    raw_rfp_file.columns = raw_rfp_file.iloc[3]
    raw_rfp_file=raw_rfp_file[3:]
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Proposal Sheet']
    
    red_pattern = PatternFill(patternType='solid',fgColor='00FF0000')
    yellow_pattern = PatternFill(patternType='solid',fgColor='00FFFF00')
    
    
    Total_Delivered_Cost_non_valid_numeric=[]
    Non_valid_currency=[]
    Non_valid_bid_price_base=[]
    Non_valid_country_warning=[]
    Non_valid_Expected_Lead_Time_warning=[]
    Non_valid_100_fulfil_demand=[]
    Non_valid_non_valid_FOB_april_warning=[]
    Non_valid_non_valid_FOB_q2_warning=[]
    Non_valid_eu_reach_hcb_limit_warning=[]
    Non_valid_functionlly_equivalent_material_warning=[]
    Non_valid_Alternate_Product_Pricing=[]

    for index in raw_rfp_file.index.tolist():
        
        if is_valid_bid_price_base(raw_rfp_file['Do you intend to bid price based on Formula or Fixed?'][index]) is False:
            if index>3:
                Non_valid_bid_price_base.append(index+2)
                index_val=str(printString(raw_rfp_file.columns.get_loc('Do you intend to bid price based on Formula or Fixed?')+1))+str(index+2).strip()
                index_val=remove_empty_space(index_val)            
                ws[index_val].fill = red_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Do you intend to bid price based on Formula or Fixed?')+1)+str(5))].fill = red_pattern
                
        if is_valid_expected_lead_time_warning(raw_rfp_file['Expected Lead Time (Days)'][index]) is False:
            if index>3:
                Non_valid_Expected_Lead_Time_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('Expected Lead Time (Days)')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Expected Lead Time (Days)')+1)+str(5))].fill = yellow_pattern

                
        if is_valid_can_you_fulfil_100_of_demand(raw_rfp_file['Can you fulfil 100% of demand? (Y/N)'][index]) is False:
            if index>3:
                Non_valid_100_fulfil_demand.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('Can you fulfil 100% of demand? (Y/N)')+1)+str(index+2)))].fill = red_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Can you fulfil 100% of demand? (Y/N)')+1)+str(5))].fill = red_pattern
                
        if is_valid_numeric_warning(raw_rfp_file['FOB Price based on Formula/Fixed April, 2023'][index]) is False:
            if index>3:
                Non_valid_non_valid_FOB_april_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('FOB Price based on Formula/Fixed April, 2023')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('FOB Price based on Formula/Fixed April, 2023')+1)+str(5))].fill = yellow_pattern


        if is_valid_numeric_warning(raw_rfp_file['FOB Price based on Formula/Fixed Q2 2023'][index]) is False:
            if index>3:
                Non_valid_non_valid_FOB_q2_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('FOB Price based on Formula/Fixed Q2 2023')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('FOB Price based on Formula/Fixed Q2 2023')+1)+str(5))].fill = yellow_pattern
                
        if is_valid_numeric(raw_rfp_file['Total Delivered Cost'][index]) is False:
            if index>3:
                Total_Delivered_Cost_non_valid_numeric.append(index+2)
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Total Delivered Cost')+1)+str(index+2))].fill = red_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Total Delivered Cost')+1)+str(5))].fill = red_pattern
        
        if is_valid_currency(raw_rfp_file["In case quote has been calcuated in a different currency please provide exchange rate to 'Preferred Currency'"][index]) is False:
            if index>3:
                Non_valid_currency.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc("In case quote has been calcuated in a different currency please provide exchange rate to 'Preferred Currency'")+1))+str(index+2))].fill = red_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc("In case quote has been calcuated in a different currency please provide exchange rate to 'Preferred Currency'")+1)+str(5))].fill = red_pattern

        if is_valid_hcb_limit_warning(raw_rfp_file['Is Material Compatible with EU REACH and new HCB limits?'][index]) is False:
            if index>3:
                Non_valid_eu_reach_hcb_limit_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('Is Material Compatible with EU REACH and new HCB limits?')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Is Material Compatible with EU REACH and new HCB limits?')+1)+str(5))].fill = yellow_pattern
                
        if is_valid_hcb_limit_warning(raw_rfp_file['Alternate/ Functionally equivalent material'][index]) is False:
            if index>3:
                Non_valid_functionlly_equivalent_material_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('Alternate/ Functionally equivalent material')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Alternate/ Functionally equivalent material')+1)+str(5))].fill = yellow_pattern
                

        if is_valid_numeric_warning(raw_rfp_file['Alternate Product Pricing'][index]) is False:
            if index>3:
                Non_valid_Alternate_Product_Pricing.append(index+2)
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc("Alternate Product Pricing"+1))+str(index+2))].fill = red_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc("Alternate Product Pricing")+1)+str(5))].fill = red_pattern

        if is_valid_country(raw_rfp_file['Country of Origin'][index]) is False:
            
            if index>3:
                Non_valid_country_warning.append(index+2)
                ws[remove_empty_space(str(printString(raw_rfp_file.columns.get_loc('Country of Origin')+1)+str(index+2)))].fill = yellow_pattern
                ws[remove_empty_space(printString(raw_rfp_file.columns.get_loc('Country of Origin')+1)+str(5))].fill = yellow_pattern
                
    
    error={}
    error['Total Delivered Cost']=Total_Delivered_Cost_non_valid_numeric
    error["In case quote has been calcuated in a different currency please provide exchange rate to 'Preferred Currency'"]=Non_valid_currency
    error['Do you intend to bid price based on Formula or Fixed?']=Non_valid_bid_price_base
    error['Can you fulfil 100% of demand? (Y/N)']=Non_valid_100_fulfil_demand
    
    warning={}
    warning['Expected Lead Time (Days)']=Non_valid_Expected_Lead_Time_warning
    warning['Country of Origin']=Non_valid_country_warning
    warning['FOB Price based on Formula/Fixed April, 2023']=Non_valid_non_valid_FOB_april_warning
    warning['FOB Price based on Formula/Fixed Q2 2023']=Non_valid_non_valid_FOB_q2_warning
    warning['Is Material Compatible with EU REACH and new HCB limits?']=Non_valid_eu_reach_hcb_limit_warning
    warning['Alternate/ Functionally equivalent material']=Non_valid_functionlly_equivalent_material_warning
    warning['Alternate Product Pricing']=Non_valid_Alternate_Product_Pricing
    
    
    
    error_flag=False
    warning_flag=False
    listt_errors=[Total_Delivered_Cost_non_valid_numeric,Non_valid_currency,Non_valid_bid_price_base,Non_valid_100_fulfil_demand]

    list_warning=[
    Non_valid_country_warning,
    Non_valid_Expected_Lead_Time_warning,
    Non_valid_non_valid_FOB_april_warning,
    Non_valid_non_valid_FOB_q2_warning,
    Non_valid_eu_reach_hcb_limit_warning,
    Non_valid_functionlly_equivalent_material_warning,
    Non_valid_Alternate_Product_Pricing
    ]
    
    for listtt in list_warning:
        if len(listtt)>0:
            warning_flag=True
            break
    

    for listtt in listt_errors:
        if len(listtt)>0:
            error_flag=True
            break
        
    print(error)
    
    if error_flag==True:

        wb.save(f"file/RFP_colored_{start_timestamp}.xlsx")
        return error_flag,error,warning_flag,warning ,raw_rfp_file,f"file/RFP_colored_{start_timestamp}.xlsx"
    
    return error_flag,error,warning_flag,warning ,raw_rfp_file,None


# In[9]:


# 1. file structure
#             - Check if input file structure is valid (E)

def processing_spend_base(file_path):
    try:
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb['Solvents']

        red_pattern = PatternFill(patternType='solid',fgColor='00FF0000')
        yellow_pattern = PatternFill(patternType='solid',fgColor='00FFFF00')

        xls = pd.ExcelFile(file_path)
        false_file_structure=0
        raw_spend_file = pd.read_excel(xls, 'Solvents')

        required_Cols=['Year', 'PostingMonth', 'Region', 'LEVEL0','LEVEL1', 'LEVEL2', 'LEVEL3', 'VendorParent','VendorName','Vendor Region', 'MaterialDesc', 'Delivery type', 'TipsCode', 'TipsDesc',
               'Purchasingdoc','PaymentTerms', 'BASE_UOM', 'PoCurrency', 'LOC_CURRCY',
               'CALQUANTITY', 'QuantityKG', 'AmountLC', 'AmountInUSD', 'UnitPrice']


        for col in required_Cols:
            if col not in raw_spend_file.columns:
                false_file_structure=1

        Non_valid_year=[]
        Non_valid_posting_month=[]
        Non_valid_tips_code=[]
        Non_valid_delivery_type=[]
        Non_valid_amount_in_usd=[]
        Non_valid_amount_lc=[]
        Non_valid_qunatity_kg=[]
        Non_valid_cal_quantity=[]
        Non_valid_unit_price=[]

        for index in raw_spend_file.index.tolist():
                if is_valid_alternate_product_pricing(raw_spend_file['Year'][index]) is False:
                    Non_valid_year.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("Year")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("Year")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['PostingMonth'][index]) is False:
                    Non_valid_posting_month.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("PostingMonth")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("PostingMonth")+1)+str(1))].fill = red_pattern


                if is_not_null(raw_spend_file['TipsCode'][index]) is False:
                    Non_valid_tips_code.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("TipsCode")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("TipsCode")+1)+str(1))].fill = red_pattern


                if is_only_char_string(raw_spend_file['Delivery type'][index]) is False:
                    Non_valid_delivery_type.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("Delivery type")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("Delivery type")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['UnitPrice'][index]) is False or is_not_null(raw_spend_file['UnitPrice'][index]) is False:
                    Non_valid_unit_price.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("UnitPrice")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("UnitPrice")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['CALQUANTITY'][index]) is False or is_not_null(raw_spend_file['CALQUANTITY'][index]) is False:
                    Non_valid_cal_quantity.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("CALQUANTITY")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("CALQUANTITY")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['QuantityKG'][index]) is False or is_not_null(raw_spend_file['QuantityKG'][index]) is False:
                    Non_valid_qunatity_kg.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("QuantityKG")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("QuantityKG")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['AmountLC'][index]) is False or is_not_null(raw_spend_file['AmountLC'][index]) is False:
                    Non_valid_amount_lc.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("AmountLC")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("AmountLC")+1)+str(1))].fill = red_pattern


                if is_valid_alternate_product_pricing(raw_spend_file['AmountInUSD'][index]) is False or is_not_null(raw_spend_file['AmountInUSD'][index]) is False:
                    Non_valid_amount_in_usd.append(index)
                    ws[remove_empty_space(str(printString(raw_spend_file.columns.get_loc("AmountInUSD")+1))+str(index+2))].fill = red_pattern
                    ws[remove_empty_space(printString(raw_spend_file.columns.get_loc("AmountInUSD")+1)+str(1))].fill = red_pattern

        
        error={}
        error['Year']=Non_valid_year
        error["PostingMonth"]=Non_valid_posting_month
        error['TipsCode']=Non_valid_tips_code
        error['Delivery type']=Non_valid_delivery_type
        error['UnitPrice']=Non_valid_unit_price
        error['CALQUANTITY']=Non_valid_cal_quantity    
        error['QuantityKG']=Non_valid_qunatity_kg    
        error['AmountLC']=Non_valid_amount_lc    
        error['AmountInUSD']=Non_valid_amount_in_usd    

        error_flag=False
        listt_errors=[Non_valid_amount_in_usd,Non_valid_amount_lc,Non_valid_qunatity_kg,Non_valid_cal_quantity,Non_valid_unit_price,Non_valid_tips_code,Non_valid_posting_month,Non_valid_year,Non_valid_delivery_type]

        for listtt in listt_errors:
            if len(listtt)>0:
                error_flag=True
                break
        print(error_flag,error)
        
        
        if error_flag==True:
            wb.save(f"file/SPEND_colored_{start_timestamp}.xlsx")
            return error_flag,error ,raw_spend_file,f"file/SPEND_colored_{start_timestamp}.xlsx"
        return error_flag,error ,raw_spend_file,None
    except:
        return None
    

# In[11]:

@app.get("/validation")
async def main(project_id,file_type_id,file_name):    

    file_path=f'file/filepath1{start_timestamp}.xlsx' #local file path to download from adls
    
    Adls_target_folder=get_adls_folder(file_type_id)
    print('Adls folder set')

    is_error=None
    is_warning=None
    raw_rfp_file=None
    error=None
    error_file_path=None

    #get file id
    file_df=new_dbobj.read_table(f"select id from chempulse_rfp.tbl_file_info tfi where project_id = {project_id} and file_type_id = {file_type_id} and file_name like '{file_name}';")
    file_id=int(file_df['id'])

    print("file_id",file_id)

    query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, file_info_id) values(2,1,'{file_id}')"
    new_dbobj.execute_query(query_str)

    print("State :",2)
    #create if folder doesn't exist
    create_local_folder()
    read_file_adls(file_name,file_path,Adls_target_folder)

    error_cols=[]
    warning_cols=[]
    column_error={}
    column_warning={}
    
    #Validation Call
    if int(str(file_type_id).strip())==3:#RFP
        is_error, error, is_warning, warning, raw_rfp_file,error_file_path=processing(file_path)

        error_cols=list(error.keys())
        warning_cols=list(warning.keys())

        column_error={"Total Delivered Cost":'Non numeric data',
            "In case quote has been calcuated in a different currency please provide exchange rate to 'Preferred Currency'":"Invalid Currency Type Found",
            "Do you intend to bid price based on Formula or Fixed?":"Invalid Formula Type Found",
            'Can you fulfil 100% of demand? (Y/N)' : "Invalid numerical or (Yes/No) response not found"
        }


        column_warning={
            "Country of Origin":"No Actual Country or Country Code Found",
            'Expected Lead Time (Days)':"No Valid Month or Days Specified",
            'FOB Price based on Formula/Fixed April, 2023':"No Valid numeric data found",
            'FOB Price based on Formula/Fixed Q2 2023':"No Valid numeric data found",
            'Is Material Compatible with EU REACH and new HCB limits?':"Non Valid Values Other than Boolean or None",      
            'Alternate/ Functionally equivalent material':"Non Valid Values Other than Boolean or None",
            'Alternate Product Pricing': "No Valid numeric data found"
        }
    
    elif int(str(file_type_id).strip())==2:
        
        is_error, error, raw_rfp_file,error_file_path =processing_spend_base(file_path)

        print("is_error",is_error,"\nerror",error,"\nerror_file_path",error_file_path)

        error_cols=list(error.keys())
        
        column_error={
        "Year":"No Actual Numeric or Null Value Found",
        'PostingMonth':"No Actual Numeric or Null Value Found",
        'TipsCode':"No Actual Alphanumeric or Null Value Found",
        'Delivery type':"No Valid String found",
        'UnitPrice':"No Actual Numeric or Null Value Found",      
        'CALQUANTITY':"No Actual Numeric or Null Value Found",
        'AmountLC': "No Actual Numeric or Null Value Found",
        'AmountInUSD': "No Actual Numeric or Null Value Found",
        'QuantityKG':"No Actual Numeric or Null Value Found"
        }


    if is_error:
        for col in error_cols:
            if len(error[col])>0:
                replaced_col=col.replace("'","''")            
                query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, column_name, Error_type, Error, file_info_id) values(3,0,'{replaced_col}','{column_error[col]}','{error[col]}',{file_id})"
                new_dbobj.execute_query(query_str)
        
        for col in warning_cols:
            if len(warning[col])>0:
                replaced_col=col.replace("'","''")            
                query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, column_name, Error_type, Error, file_info_id) values(3,0,'{replaced_col}','{column_warning[col]}','{warning[col]}',{file_id})"
                new_dbobj.execute_query(query_str)
        
        # print(error_file_path)
        # print(file_name)
        upload_excel(os.path.join(pathlib.Path().resolve(), error_file_path),_connectionString, _containerName, os.path.join(Adls_target_folder,'ErrorFile',file_name))
        os.remove(error_file_path)
    
    else:        
        if is_warning:
            for col in warning_cols:
                if len(warning[col])>0:
                    replaced_col=col.replace("'","''")            
                    query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, column_name, Error_type, Error, file_info_id) values(3,2,'{replaced_col}','{column_warning[col]}','{warning[col]}',{file_id})"
                    new_dbobj.execute_query(query_str)
        else:
            query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, file_info_id) values(3,1,{file_id})"
            new_dbobj.execute_query(query_str)
        
        #Standardization of columns
        
        if file_type_id==3:
            raw_rfp_file['Can you fulfil 100% of demand? (Y/N)']=raw_rfp_file['Can you fulfil 100% of demand? (Y/N)'].apply(update_Can_you_fulfil_100_demand)
            raw_rfp_file['Expected Lead Time (Days)']=raw_rfp_file['Expected Lead Time (Days)'].apply(standardize_Expected_Lead_Time)
        
        query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, file_info_id) values(4,1,{file_id})"
        new_dbobj.execute_query(query_str)

        os.path.join(Adls_target_folder,'Output',file_name)
        upload_excel(os.path.join(pathlib.Path().resolve(), file_path),_connectionString, _containerName, os.path.join(Adls_target_folder,'Output',file_name))

        query_str=f"insert into chempulse_rfp.tbl_file_log(file_state_id, state_successful, file_info_id) values(5,1,{file_id})"
        new_dbobj.execute_query(query_str)
        
    os.remove(os.path.join(pathlib.Path().resolve(), file_path))
    

if __name__=='__main__':
    uvicorn.run(app,port=8080)

